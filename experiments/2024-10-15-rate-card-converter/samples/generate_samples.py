"""Script to generate synthetic freight carrier rate card sample files.

Produces two files for manual testing and one pair of minimal files for
the automated test suite:

Sample files (realistic carrier data, written to --output-dir):
  sample_rate_card.xlsx - FastFreight Logistics, USD per pound (requires conversion),
                          3 sheets: Rates, Surcharges, Info metadata.
  sample_rate_card.pdf  - FreshCargo GmbH, EUR per kg, text-based PDF.

Test fixtures (minimal data, written to tests/fixtures/ when --test-fixtures):
  tests/fixtures/sample.xlsx - Minimal 3-row Excel file for unit tests.
  tests/fixtures/sample.pdf  - Minimal text PDF for unit tests.

The PDF is generated using raw PDF syntax (no third-party PDF library required).
A minimal valid PDF is a sequence of objects with a content stream. The content
is text rendered in Helvetica at fixed positions using PDF content stream operators.
pdfplumber can parse these files since they use standard PDF text stream format.

Usage:
    python samples/generate_samples.py
    python samples/generate_samples.py --test-fixtures
    python samples/generate_samples.py --output-dir /custom/path --test-fixtures
"""

from __future__ import annotations

import argparse
import struct
import textwrap
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def build_sample_xlsx(output_path: Path) -> None:
    """Generate a realistic synthetic Excel rate card with non-standard headers.

    The column headers are deliberately non-standard (carrier-specific naming,
    weights in pounds, rates in USD/lb) to demonstrate the LLM's ability to
    map them to the canonical schema.

    Args:
        output_path: Destination file path for the generated .xlsx file.
    """
    wb = openpyxl.Workbook()

    _build_rates_sheet(wb.active)
    wb.active.title = "Rates"

    _build_surcharges_sheet(wb.create_sheet("Surcharges"))
    _build_info_sheet(wb.create_sheet("Info"))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Generated: {output_path}")


def _build_rates_sheet(ws: object) -> None:
    """Populate the Rates worksheet with non-standard headers and realistic data.

    Headers use carrier-specific naming that must be semantically mapped:
    - 'From Zone' -> origin
    - 'To Zone' -> destination
    - 'Service Type' -> service_level (with non-standard values like 'Priority Air')
    - 'Min Weight (lbs)' / 'Max Weight (lbs)' -> weight_min_kg / weight_max_kg (with conversion)
    - 'Price/lb (USD)' -> rate_per_kg (with unit conversion)
    - 'Min Charge' -> min_charge
    - 'FSC %' -> fuel_surcharge_pct

    Args:
        ws: openpyxl Worksheet to populate.
    """
    headers = [
        "From Zone", "To Zone", "Service Type",
        "Min Weight (lbs)", "Max Weight (lbs)", "Price/lb (USD)",
        "Min Charge", "FSC %", "Valid From", "Valid To",
    ]

    # Style header row.
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Synthetic rate data. Weights are in pounds; rates are USD per pound.
    # The LLM should convert to kg (multiply lb by 0.453592) and per-kg rate.
    rate_data = [
        ["US-West", "Asia-Pacific",  "Priority Air",   1.1,  22.0, 5.50, 18.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-West", "Asia-Pacific",  "Economy Ground", 22.1, 110.0, 2.80, 12.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-East", "Europe",        "Priority Air",    1.1,  22.0, 4.90, 15.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-East", "Europe",        "Standard",       22.1, 220.0, 2.20, 10.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-Central", "Latin Am",   "Economy Ground", 1.1,  55.0, 3.10, 12.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-Central", "Latin Am",   "Priority Air",   1.1,  22.0, 5.20, 16.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-West", "Australia",     "Standard",       22.1, 110.0, 3.40, 14.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-East", "Middle East",   "Priority Air",    1.1,  22.0, 6.10, 20.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-West", "Southeast Asia","Standard",       22.1, 220.0, 2.60, 11.00, 5.5, "2024-10-01", "2024-12-31"],
        ["US-East", "Africa",        "Economy Ground", 1.1,  55.0, 3.80, 13.00, 5.5, "2024-10-01", "2024-12-31"],
    ]

    for row_data in rate_data:
        ws.append(row_data)

    # Add carrier info in a cell outside the table to test metadata extraction.
    ws["L1"] = "Carrier: FastFreight Logistics"
    ws["L2"] = "Contact: rates@fastfreight.com"


def _build_surcharges_sheet(ws: object) -> None:
    """Populate the Surcharges worksheet.

    Args:
        ws: openpyxl Worksheet to populate.
    """
    ws.append(["Surcharge Name", "Pct", "Flat Fee (USD)", "Conditions"])
    ws.append(["Fuel Surcharge", 5.5, None, "Applied to all shipments"])
    ws.append(["Peak Season", 8.0, None, "Nov 15 - Jan 5 only"])
    ws.append(["Residential Delivery", None, 4.50, "Residential addresses only"])
    ws.append(["Dangerous Goods", 15.0, None, "IATA DG Class 1-9"])


def _build_info_sheet(ws: object) -> None:
    """Populate the Info worksheet with unstructured metadata.

    Args:
        ws: openpyxl Worksheet to populate.
    """
    ws["A1"] = "Rate Card Information"
    ws["A2"] = "Carrier Name:"
    ws["B2"] = "FastFreight Logistics"
    ws["A3"] = "Carrier Code:"
    ws["B3"] = "FFL"
    ws["A4"] = "Currency:"
    ws["B4"] = "USD"
    ws["A5"] = "Weight Unit:"
    ws["B5"] = "lbs"
    ws["A6"] = "Rate Unit:"
    ws["B6"] = "per lb"
    ws["A7"] = "Effective From:"
    ws["B7"] = "2024-10-01"
    ws["A8"] = "Effective To:"
    ws["B8"] = "2024-12-31"
    ws["A9"] = "Notes:"
    ws["B9"] = "All rates are exclusive of taxes and duties."


def build_minimal_xlsx(output_path: Path) -> None:
    """Generate a minimal Excel file for unit tests.

    Contains one sheet with 3 data rows and standard headers to verify basic
    extraction functionality without LLM involvement.

    Args:
        output_path: Destination file path.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rates"
    ws.append(["From", "To", "Service", "Rate USD/kg", "Min KG", "Max KG"])
    ws.append(["US", "EU", "express", 12.50, 0.5, 10.0])
    ws.append(["US", "APAC", "standard", 8.00, 1.0, 30.0])
    ws.append(["EU", "US", "economy", 6.50, 5.0, 50.0])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Generated: {output_path}")


# ---------------------------------------------------------------------------
# PDF generation (raw PDF syntax — no third-party library required)
# ---------------------------------------------------------------------------

def build_sample_pdf(output_path: Path) -> None:
    """Generate a text-based PDF rate card using raw PDF content stream syntax.

    The PDF contains a single page with a text table of freight rates.
    Column headers differ from the Excel sample to test the LLM's carrier-
    agnostic field mapping. Rates are in EUR per kg (no unit conversion needed).
    Footer text provides carrier name for metadata extraction testing.

    pdfplumber can parse this PDF because it uses standard Type1 font references
    and text rendering operators (BT/ET, Tf, Td, Tj).

    Args:
        output_path: Destination file path for the generated .pdf file.
    """
    content_lines = _build_pdf_content_lines()
    pdf_bytes = _assemble_pdf(content_lines)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(pdf_bytes)
    print(f"Generated: {output_path}")


def _build_pdf_content_lines() -> list[str]:
    """Build the list of text lines to render in the PDF content stream.

    Returns:
        A list of strings, each to be rendered as a line in the PDF.
    """
    lines = [
        "FreshCargo GmbH - Freight Rate Card Q4 2024",
        "",
        "Origin       | Destination  | Svc      | Min KG | Max KG | Rate EUR/kg",
        "-------------|--------------|----------|--------|--------|------------",
        "DE-Frankfurt | US-New York  | express  | 0.50   | 10.00  | 9.80",
        "DE-Frankfurt | US-New York  | standard | 10.10  | 50.00  | 5.60",
        "DE-Frankfurt | US-New York  | economy  | 50.10  | 200.00 | 3.20",
        "DE-Hamburg   | CN-Shanghai  | express  | 0.50   | 15.00  | 11.50",
        "DE-Hamburg   | CN-Shanghai  | standard | 15.10  | 100.00 | 6.80",
        "DE-Munich    | JP-Tokyo     | express  | 0.50   | 10.00  | 12.10",
        "DE-Munich    | JP-Tokyo     | standard | 10.10  | 50.00  | 7.20",
        "NL-Amsterdam | AU-Sydney    | economy  | 5.00   | 100.00 | 4.50",
        "FR-Paris     | BR-Sao Paulo | standard | 1.00   | 30.00  | 8.90",
        "FR-Paris     | ZA-Cape Town | economy  | 5.00   | 50.00  | 5.10",
        "",
        "Surcharges: Fuel 6.0% | Peak Season Nov-Jan 9.0%",
        "Currency: EUR | Weight Unit: kg",
        "Valid: 2024-10-01 to 2024-12-31",
        "",
        "FreshCargo GmbH | Rate Card Q4 2024 | Confidential",
    ]
    return lines


def _assemble_pdf(text_lines: list[str]) -> bytes:
    """Assemble a minimal valid PDF document from a list of text lines.

    Uses raw PDF object syntax with a BT...ET content stream. Each line is
    rendered using the Td operator to advance to the next line. The font is
    Helvetica (a standard Type1 font, no embedding required).

    Args:
        text_lines: Lines of text to render in the PDF.

    Returns:
        Raw bytes of a valid PDF document.
    """
    # Build the text content stream. Each line uses Td to move down by 14 points.
    content_parts = ["BT", "/F1 9 Tf", "40 780 Td", "14 TL"]
    for line in text_lines:
        escaped = line.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        content_parts.append(f"({escaped}) Tj T*")
    content_parts.append("ET")
    content_stream = "\n".join(content_parts).encode("latin-1")

    # Object 1: Catalog
    obj1 = b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
    # Object 2: Pages
    obj2 = b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
    # Object 3: Page
    obj3 = (
        b"3 0 obj\n"
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 842]\n"
        b"   /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\n"
        b"endobj\n"
    )
    # Object 4: Content stream
    stream_length = len(content_stream)
    obj4 = (
        f"4 0 obj\n<< /Length {stream_length} >>\nstream\n".encode("latin-1")
        + content_stream
        + b"\nendstream\nendobj\n"
    )
    # Object 5: Font
    obj5 = (
        b"5 0 obj\n"
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica "
        b"/Encoding /WinAnsiEncoding >>\nendobj\n"
    )

    header = b"%PDF-1.4\n"
    body_parts = [obj1, obj2, obj3, obj4, obj5]
    offsets: list[int] = []
    body = b""
    pos = len(header)
    for part in body_parts:
        offsets.append(pos)
        body += part
        pos += len(part)

    xref_offset = len(header) + len(body)
    xref = b"xref\n"
    xref += f"0 {len(body_parts) + 1}\n".encode()
    xref += b"0000000000 65535 f \n"
    for offset in offsets:
        xref += f"{offset:010d} 00000 n \n".encode()

    trailer = (
        f"trailer\n<< /Size {len(body_parts) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF\n"
    ).encode("latin-1")

    return header + body + xref + trailer


def build_minimal_pdf(output_path: Path) -> None:
    """Generate a minimal valid PDF for unit tests.

    Contains a single page with a simple 3-row rate table as text.

    Args:
        output_path: Destination file path.
    """
    lines = [
        "Test Rate Card",
        "",
        "Origin | Destination | Service  | Rate    | Currency",
        "-------|-------------|----------|---------|--------",
        "US     | EU          | express  | 10.50   | USD",
        "US     | APAC        | standard | 7.80    | USD",
    ]
    pdf_bytes = _assemble_pdf(lines)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(pdf_bytes)
    print(f"Generated: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    """Parse arguments and generate the requested sample files."""
    parser = argparse.ArgumentParser(
        description="Generate synthetic rate card sample files for testing."
    )
    parser.add_argument(
        "--output-dir",
        default="samples",
        help="Directory for full sample files (default: samples/).",
    )
    parser.add_argument(
        "--test-fixtures",
        action="store_true",
        help="Also generate minimal test fixtures in tests/fixtures/.",
    )
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    build_sample_xlsx(output_dir / "sample_rate_card.xlsx")
    build_sample_pdf(output_dir / "sample_rate_card.pdf")

    if args.test_fixtures:
        fixtures_dir = Path("tests/fixtures")
        build_minimal_xlsx(fixtures_dir / "sample.xlsx")
        build_minimal_pdf(fixtures_dir / "sample.pdf")
        print("Test fixtures written to tests/fixtures/.")


if __name__ == "__main__":
    main()
