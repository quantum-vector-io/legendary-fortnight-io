"""Excel document extractor using openpyxl for local and test environments.

openpyxl is a synchronous library. All parsing runs inside asyncio.to_thread()
to avoid blocking the event loop, consistent with the PDF extractor pattern.

data_only=True is critical: without it, openpyxl returns formula strings
(e.g. '=SUM(A1:A5)') rather than computed values. read_only=True reduces
memory consumption for large workbooks by streaming rows.
"""

from __future__ import annotations

import asyncio
import io
from typing import Any

import openpyxl

from src.domain.exceptions import DocumentExtractionError, UnsupportedFileFormatError
from src.domain.ports.document_extractor import DocumentExtractorPort, ExtractedTable

# XLSX files are ZIP archives. All ZIP files start with the PK\x03\x04 magic bytes.
_XLSX_MAGIC_BYTES = b"PK\x03\x04"


class XLSXExtractor(DocumentExtractorPort):
    """Extracts tabular data from Excel (.xlsx) documents using openpyxl.

    Each worksheet in the workbook is extracted as a separate ExtractedTable.
    Leading empty rows are skipped, and the first non-empty row is treated as
    column headers. Rows where all cells are empty are filtered out.
    """

    async def extract(self, file_bytes: bytes, filename: str) -> list[ExtractedTable]:
        """Extract tabular data from an Excel document.

        Args:
            file_bytes: Raw Excel document bytes.
            filename: Original filename for logging and error context.

        Returns:
            A list of ExtractedTable instances, one per worksheet.

        Raises:
            UnsupportedFileFormatError: If file_bytes is not a valid XLSX file.
            DocumentExtractionError: If openpyxl fails to parse the document.
        """
        return await asyncio.to_thread(self._extract_sync, file_bytes, filename)

    def _extract_sync(self, file_bytes: bytes, filename: str) -> list[ExtractedTable]:
        """Synchronous Excel extraction logic for thread pool execution.

        Args:
            file_bytes: Raw Excel bytes.
            filename: Filename for error context.

        Returns:
            List of ExtractedTable instances.

        Raises:
            UnsupportedFileFormatError: If format validation fails.
            DocumentExtractionError: If openpyxl parsing fails.
        """
        self._validate_magic_bytes(file_bytes, filename)
        try:
            return self._parse_workbook(file_bytes)
        except (UnsupportedFileFormatError, DocumentExtractionError):
            raise
        except Exception as exc:
            raise DocumentExtractionError(
                f"Failed to parse Excel file '{filename}': {exc}",
                detail={"filename": filename, "error": str(exc)},
            ) from exc

    def _validate_magic_bytes(self, file_bytes: bytes, filename: str) -> None:
        """Verify the file begins with the ZIP/XLSX magic byte sequence.

        Args:
            file_bytes: File bytes to check.
            filename: Filename for error message context.

        Raises:
            UnsupportedFileFormatError: If the magic bytes are absent.
        """
        if file_bytes[:4] != _XLSX_MAGIC_BYTES:
            raise UnsupportedFileFormatError(
                f"File '{filename}' does not appear to be a valid XLSX file.",
                detail={"filename": filename},
            )

    def _parse_workbook(self, file_bytes: bytes) -> list[ExtractedTable]:
        """Open the workbook and extract one table per worksheet.

        Args:
            file_bytes: Validated XLSX bytes.

        Returns:
            List of ExtractedTable instances.
        """
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )
        results: list[ExtractedTable] = []
        for sheet in workbook.worksheets:
            table = self._extract_sheet(sheet)
            if table is not None and not table.is_empty():
                results.append(table)
        workbook.close()
        return results

    def _extract_sheet(self, sheet: Any) -> ExtractedTable | None:
        """Extract an ExtractedTable from a single openpyxl worksheet.

        Scans rows top-to-bottom, skipping leading empty rows. The first
        non-empty row becomes the headers list.

        Args:
            sheet: An openpyxl worksheet (read_only or standard).

        Returns:
            An ExtractedTable with headers and data rows, or None if the
            sheet has no non-empty rows.
        """
        all_rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            normalized = self._normalize_row(row)
            all_rows.append(normalized)

        # Skip leading empty rows.
        start_index = self._find_first_data_row(all_rows)
        if start_index is None:
            return None

        data_rows = all_rows[start_index:]
        if not data_rows:
            return None

        headers = data_rows[0]
        rows = [
            row for row in data_rows[1:]
            if any(cell for cell in row)
        ]

        return ExtractedTable(
            headers=headers,
            rows=rows,
            sheet_name=sheet.title,
        )

    def _normalize_row(self, row: tuple[Any, ...]) -> list[str]:
        """Convert an openpyxl row tuple to a list of strings.

        None values (empty cells) are converted to empty strings. All other
        values are converted with str() and whitespace-stripped.

        Args:
            row: A tuple of cell values from openpyxl iter_rows(values_only=True).

        Returns:
            A list of string cell values.
        """
        return [str(cell).strip() if cell is not None else "" for cell in row]

    def _find_first_data_row(self, rows: list[list[str]]) -> int | None:
        """Find the index of the first row that contains at least one non-empty cell.

        Args:
            rows: All rows from the sheet as normalized string lists.

        Returns:
            The 0-based index of the first non-empty row, or None if all rows
            are empty.
        """
        for index, row in enumerate(rows):
            if any(cell for cell in row):
                return index
        return None
